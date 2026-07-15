"""Creazione e lettura/scrittura del file Excel condiviso.

Fogli:
  Prenotazioni — matrice evento x membro: metti una X per prenotarti
  Alloggi      — risultati Booking/Airbnb scritti dalla pipeline
  Voli         — risultati Google Flights scritti dalla pipeline
  Eventi       — calendario di riferimento (generato da data/events_2027.yaml)
"""
from __future__ import annotations

from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .config import settings
from .models import Event, FlightOption, StayOption, TripRequest
from .registry import load_events, load_team

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INTL_FILL = PatternFill("solid", fgColor="FCE4D6")
WARN_FONT = Font(color="9C5700", italic=True)

BOOKINGS_SHEET = "Prenotazioni"
STAYS_SHEET = "Alloggi"
FLIGHTS_SHEET = "Voli"
EVENTS_SHEET = "Eventi"

# Colonne fisse del foglio Prenotazioni prima dei nomi dei membri
FIXED_COLS = ["Evento", "Tipo", "Città", "Date evento", "Check-in", "Check-out"]


def create_workbook(path=None) -> str:
    path = path or settings.workbook_path
    events = load_events()
    members, _ = load_team()
    names = [m["name"] for m in members]

    wb = Workbook()
    _build_bookings_sheet(wb.active, events, names)
    _build_results_sheets(wb)
    _build_events_sheet(wb.create_sheet(EVENTS_SHEET), events)

    wb.save(path)
    return str(path)


def _build_bookings_sheet(ws, events: list[Event], names: list[str]) -> None:
    ws.title = BOOKINGS_SHEET
    headers = FIXED_COLS + names + ["Tot persone"]
    ws.append(headers)
    _style_header(ws, len(headers))

    first_member_col = len(FIXED_COLS) + 1
    last_member_col = len(FIXED_COLS) + len(names)
    dv = DataValidation(type="list", formula1='"X"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)

    for row_idx, ev in enumerate(events, start=2):
        ws.cell(row_idx, 1, ev.name)
        ws.cell(row_idx, 2, _type_label(ev.type))
        ws.cell(row_idx, 3, f"{ev.city} ({ev.country})")
        ws.cell(row_idx, 4, f"{_fmt(ev.start)} – {_fmt(ev.end)}")
        ws.cell(row_idx, 5, ev.default_check_in).number_format = "DD/MM/YYYY"
        ws.cell(row_idx, 6, ev.default_check_out).number_format = "DD/MM/YYYY"
        col_a = get_column_letter(first_member_col)
        col_b = get_column_letter(last_member_col)
        total = ws.cell(row_idx, last_member_col + 1, f'=COUNTIF({col_a}{row_idx}:{col_b}{row_idx},"X")')
        total.font = Font(bold=True)
        for col in range(first_member_col, last_member_col + 1):
            cell = ws.cell(row_idx, col)
            cell.alignment = Alignment(horizontal="center")
            dv.add(cell)
        if ev.type == "international":
            for col in range(1, len(FIXED_COLS) + 1):
                ws.cell(row_idx, col).fill = INTL_FILL

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 22
    for col in ("E", "F"):
        ws.column_dimensions[col].width = 12
    for col in range(first_member_col, last_member_col + 2):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = "B2"


def _build_results_sheets(wb: Workbook) -> None:
    stays = wb.create_sheet(STAYS_SHEET)
    stays.append(
        ["Evento", "Fonte", "Nome", "Tipo", "Persone", "Notti", "Prezzo tot", "€/notte",
         "Recensioni", "N. recensioni", "Distanza fiera (km)", "Cancellazione gratuita", "Link"]
    )
    _style_header(stays, 13)
    widths = [40, 12, 42, 16, 9, 7, 11, 9, 11, 13, 18, 20, 50]
    for i, w in enumerate(widths, start=1):
        stays.column_dimensions[get_column_letter(i)].width = w
    stays.freeze_panes = "A2"

    flights = wb.create_sheet(FLIGHTS_SHEET)
    flights.append(
        ["Evento", "Da", "A", "Andata", "Ritorno", "Compagnia", "Scali", "Durata",
         "Persone", "Prezzo tot", "Link Google Flights"]
    )
    _style_header(flights, 11)
    widths = [40, 6, 6, 34, 22, 24, 7, 9, 9, 11, 60]
    for i, w in enumerate(widths, start=1):
        flights.column_dimensions[get_column_letter(i)].width = w
    flights.freeze_panes = "A2"


def _build_events_sheet(ws, events: list[Event]) -> None:
    ws.append(["Evento", "Tipo", "Città", "Paese", "Inizio", "Fine", "Sede", "Sede confermata", "Aeroporto", "Lat", "Lon"])
    _style_header(ws, 11)
    for ev in events:
        ws.append(
            [ev.name, _type_label(ev.type), ev.city, ev.country, _fmt(ev.start), _fmt(ev.end),
             ev.venue, "SÌ" if ev.venue_confirmed else "NO — da verificare", ev.airport, ev.lat, ev.lon]
        )
        if not ev.venue_confirmed:
            ws.cell(ws.max_row, 7).font = WARN_FONT
            ws.cell(ws.max_row, 8).font = WARN_FONT
    for col, w in zip("ABCDEFGHIJK", [46, 14, 16, 16, 11, 11, 30, 20, 10, 10, 10]):
        ws.column_dimensions[col].width = w


def read_trip_requests(path=None) -> list[TripRequest]:
    """Legge Prenotazioni e restituisce gli eventi con almeno una X."""
    path = path or settings.workbook_path
    wb = load_workbook(path, data_only=True)
    ws = wb[BOOKINGS_SHEET]
    events_by_name = {e.name: e for e in load_events()}
    names = [c.value for c in ws[1][len(FIXED_COLS):] if c.value and c.value != "Tot persone"]

    requests: list[TripRequest] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        event = events_by_name.get(row[0])
        if event is None:
            continue
        participants = [
            name
            for name, flag in zip(names, row[len(FIXED_COLS):])
            if isinstance(flag, str) and flag.strip().upper() == "X"
        ]
        if not participants:
            continue
        check_in = _cell_date(row[4], event.default_check_in)
        check_out = _cell_date(row[5], event.default_check_out)
        requests.append(TripRequest(event, participants, check_in, check_out))
    return requests


def write_results(
    stays: dict[str, tuple[TripRequest, list[StayOption]]],
    flights: dict[str, tuple[TripRequest, list[FlightOption]]],
    path=None,
) -> None:
    path = path or settings.workbook_path
    wb = load_workbook(path)

    ws = wb[STAYS_SHEET]
    ws.delete_rows(2, ws.max_row)
    for _, (req, options) in stays.items():
        for o in options:
            ws.append(
                [req.event.name, o.source, o.name, o.property_type, req.people, req.nights,
                 o.total_price, o.price_per_night(req.nights),
                 o.review_score if o.review_score is not None else "n.d.",
                 o.review_count if o.review_count is not None else "n.d.",
                 o.distance_km,
                 {True: "Sì", False: "No", None: "Da verificare"}[o.free_cancellation],
                 o.url]
            )

    ws = wb[FLIGHTS_SHEET]
    ws.delete_rows(2, ws.max_row)
    for _, (req, options) in flights.items():
        for f in options:
            ws.append(
                [req.event.name, f.origin, f.destination, f.outbound, f.inbound, f.airline,
                 f.stops, _fmt_duration(f.duration_min), req.people, f.price, f.url]
            )

    wb.save(path)


def write_link_results(stays: dict, flights: dict, path=None) -> None:
    """Modalità gratuita: scrive una riga per portale con il link pre-filtrato."""
    path = path or settings.workbook_path
    wb = load_workbook(path)

    ws = wb[STAYS_SHEET]
    ws.delete_rows(2, ws.max_row)
    for _, (req, links) in stays.items():
        for link in links:
            ws.append(
                [req.event.name, link.source, link.label, "Ricerca pre-filtrata",
                 req.people, req.nights, "apri link", "", "", "", "", "", link.url]
            )
            _linkify(ws.cell(ws.max_row, 13))

    ws = wb[FLIGHTS_SHEET]
    ws.delete_rows(2, ws.max_row)
    for _, (req, links) in flights.items():
        for link in links:
            ws.append(
                [req.event.name, link.origin, req.event.airport,
                 f"{_fmt(req.check_in)}", f"{_fmt(req.check_out)}",
                 link.label, "", "", req.people, "apri link", link.url]
            )
            _linkify(ws.cell(ws.max_row, 11))

    wb.save(path)


def _linkify(cell) -> None:
    cell.hyperlink = cell.value
    cell.font = Font(color="0563C1", underline="single")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _type_label(t: str) -> str:
    return {"international": "Internazionale", "regional": "Regionale", "special": "Speciale"}.get(t, t)


def _fmt(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def _fmt_duration(minutes: int) -> str:
    return f"{minutes // 60}h{minutes % 60:02d}" if minutes else "n.d."


def _cell_date(value, fallback: date) -> date:
    if isinstance(value, date):
        return value if not hasattr(value, "hour") else value.date()
    if isinstance(value, str):
        for parser in (date.fromisoformat, _parse_it_date):
            try:
                return parser(value.strip())
            except ValueError:
                pass
    return fallback


def _parse_it_date(s: str) -> date:
    day, month, year = s.split("/")
    return date(int(year), int(month), int(day))
